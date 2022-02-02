import openpyxl

fpath = r'C:\Users\o9707\Desktop\대학교\inflearn\crawling\기본편\02.네이버_주식현재가\참가자_data.xlsx'
# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트선택
ws = wb['오징어게임']

# 3) 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# 4) 엑셀 저장하기
wb.save(fpath)