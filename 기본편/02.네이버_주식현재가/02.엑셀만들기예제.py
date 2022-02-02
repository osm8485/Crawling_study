import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 4) 엑셀 저장하기
# 파일경로 지정 안할시, 가상환경 상위폴더에 결과파일저장
# 역슬래쉬\ = escape 문자, 그냥 문자열로 인식시키기 위해 문자열 앞에 'r' 붙이면 해결!
wb.save(r'C:\Users\o9707\Desktop\대학교\inflearn\crawling\기본편\02.네이버_주식현재가\참가자_data.xlsx')